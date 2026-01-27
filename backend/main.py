from fastapi import FastAPI, UploadFile, File, Form
from fastapi.responses import Response, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from sentence_transformers import SentenceTransformer, util
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from typing import List
import pdfplumber, docx, io, zipfile, os, tempfile, requests, re

class AnalysisResult(BaseModel):
    filename: str
    score: float
    status: str
    matched_keywords: List[str] = []
    missing_keywords: List[str] = []
    match_summary: str = ""

class AnalysisData(BaseModel):
    results: List[AnalysisResult]

# Force CPU usage to save memory on Render Free Tier
torch.set_num_threads(1)

app = FastAPI()

# Load model at startup to avoid memory spikes during requests
print("Loading AI Model...")
model = SentenceTransformer("all-MiniLM-L6-v2", device="cpu")
print("AI Model Loaded Successfully")

@app.get("/")
async def root():
    return {
        "message": "Resume AI Matcher API is Online",
        "status": "active",
        "model": "all-MiniLM-L6-v2"
    }

app.add_middleware(
    CORSMiddleware, 
    allow_origins=["*"], 
    allow_credentials=True, 
    allow_methods=["GET", "POST", "OPTIONS"], 
    allow_headers=["*"]
)

# Standard User-Agent for requests
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
}

def clean_text(text):
    text = text.lower()
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def get_keywords(text):
    words = set(re.findall(r'\b\w{3,}\b', text.lower()))
    stop_words = {'and', 'the', 'for', 'with', 'from', 'this', 'that', 'your', 'will', 'have', 'are', 'was', 'were', 'but', 'not', 'you', 'all', 'can', 'had', 'her', 'was', 'one', 'our', 'out', 'day', 'get', 'has', 'him', 'his', 'how', 'its', 'may', 'new', 'now', 'old', 'see', 'two', 'way', 'who', 'boy', 'did', 'use', 'her', 'she', 'try', 'why'}
    return words - stop_words

def extract_text(content, filename):
    text = ""
    try:
        ext = filename.lower()
        if ext.endswith(".pdf"):
            with pdfplumber.open(io.BytesIO(content)) as pdf: 
                text = " ".join([p.extract_text() or "" for p in pdf.pages])
        elif ext.endswith((".docx", ".doc")):
            doc = docx.Document(io.BytesIO(content))
            text = " ".join([p.text for p in doc.paragraphs])
        elif ext.endswith(".txt"): 
            text = content.decode("utf-8", errors="ignore")
    except Exception as e: 
        pass
    return text.strip()

def score_match(text, jd):
    if not text or not jd: return {"score": 0, "status": "No Content"}
    
    c_text = clean_text(text)
    c_jd = clean_text(jd)
    
    embeddings1 = model.encode(c_text, convert_to_tensor=True)
    embeddings2 = model.encode(c_jd, convert_to_tensor=True)
    semantic_score = util.cos_sim(embeddings1, embeddings2).item()
    
    jd_keywords = get_keywords(c_jd)
    resume_keywords = get_keywords(c_text)
    
    if not jd_keywords:
        keyword_score = semantic_score
    else:
        found_keywords = jd_keywords.intersection(resume_keywords)
        keyword_score = len(found_keywords) / len(jd_keywords)
    
    raw_score = (semantic_score * 0.6) + (keyword_score * 0.4)
    s = round(max(0, min(100.0, (raw_score ** 0.75) * 100 + 5)), 2)
    
    matched = list(jd_keywords.intersection(resume_keywords))
    missing = list(jd_keywords - resume_keywords)
    
    return {
        "score": s, 
        "matched_keywords": sorted(matched)[:12],
        "missing_keywords": sorted(missing)[:12],
        "match_summary": f"Matched {len(matched)} skills from JD."
    }

def apply_dynamic_statuses(results, top_n=None):
    if not results: return results
    results.sort(key=lambda x: x['score'], reverse=True)
    
    if top_n is not None and top_n > 0:
        results = results[:top_n]
    
    n = len(results)
    for i, r in enumerate(results):
        s = r['score']
        if n <= 3:
            r['status'] = "Strong" if s >= 70 else "Good" if s >= 40 else "Average" if s >= 20 else "Poor"
        else:
            if i < n * 0.25: r['status'] = "Strong"
            elif i < n * 0.50: r['status'] = "Good"
            elif i < n * 0.75: r['status'] = "Average"
            else: r['status'] = "Poor"
    return results

def get_drive_direct_link(url):
    if "docs.google.com" in url or "drive.google.com" in url:
        file_id = ""
        if "/d/" in url:
            file_id = url.split("/d/")[1].split("/")[0]
        elif "id=" in url:
            file_id = url.split("id=")[1].split("&")[0]
            
        if file_id:
            if "/spreadsheets/" in url:
                return f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
            elif "/document/" in url:
                return f"https://docs.google.com/document/d/{file_id}/export?format=docx"
            else:
                return f"https://drive.google.com/uc?export=download&id={file_id}&confirm=t"
    return url

async def process_remote_resume(url, filename, jd):
    try:
        direct_url = get_drive_direct_link(url)
        resp = requests.get(direct_url, headers=HEADERS, timeout=20)
        if resp.status_code != 200:
            return None
        if b"<!DOCTYPE html>" in resp.content[:200]:
            return None

        content_type = resp.headers.get("Content-Type", "").lower()
        ext = ".pdf"
        if "word" in content_type or ".docx" in direct_url.lower() or ".doc" in direct_url.lower():
            ext = ".docx"
        elif "text" in content_type or ".txt" in direct_url.lower():
            ext = ".txt"
        elif "pdf" in content_type or ".pdf" in direct_url.lower():
            ext = ".pdf"

        text = extract_text(resp.content, f"remote{ext}")
        if not text:
            return None
        return score_match(text, jd)
    except Exception:
        return None

@app.post("/match-folder")
async def match_folder(folder: UploadFile = File(...), job_description: str = Form(...), top_n: int = Form(default=None)):
    results = []
    try:
        content = await folder.read()
        
        # Check if it's a ZIP file
        if folder.filename.lower().endswith('.zip'):
            with tempfile.TemporaryDirectory() as tmp:
                zpath = os.path.join(tmp, "upload.zip")
                with open(zpath, "wb") as f: f.write(content)
                
                try:
                    with zipfile.ZipFile(zpath, 'r') as z:
                        z.extractall(tmp)
                except Exception as ze:
                    return JSONResponse({"error": f"Failed to unzip files: {str(ze)}"}, 400)
                
                # Process extracted files
                for root, _, files in os.walk(tmp):
                    for f in files:
                        if f.lower().endswith(('.pdf', '.docx', '.doc', '.txt')):
                            file_path = os.path.join(root, f)
                            with open(file_path, "rb") as fd:
                                d = fd.read()
                                text = extract_text(d, f)
                                if text: 
                                    results.append({"filename": f, **score_match(text, job_description)})
                        elif f.lower().endswith(('.xlsx', '.xls')):
                            # Process Excel file with resume links
                            file_path = os.path.join(root, f)
                            await process_excel_file(file_path, job_description, results)
        else:
            # Single file upload
            text = extract_text(content, folder.filename)
            if text: 
                results.append({"filename": folder.filename, **score_match(text, job_description)})
        
        results = apply_dynamic_statuses(results, top_n)
        return {"results": results}
    except Exception as e:
        return JSONResponse({"error": str(e)}, 500)

async def process_excel_file(file_path, job_description, results):
    """Process Excel file containing resume information"""
    try:
        with open(file_path, "rb") as fd:
            content = fd.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows or len(rows) < 2:
            return
            
        headers = [str(c).lower().strip() if c else "" for c in rows[0]]
        name_idx, link_idx = 0, 1
        
        for i, h in enumerate(headers):
            if any(k in h for k in ["name", "candidate", "full"]): name_idx = i
            if any(k in h for k in ["link", "resume", "drive", "url"]): link_idx = i
        
        for i, row in enumerate(rows[1:], 1):
            if not any(row): continue
            
            name = str(row[name_idx]) if name_idx < len(row) and row[name_idx] else f"Candidate_{i}"
            link = str(row[link_idx]) if link_idx < len(row) and row[link_idx] else ""
            
            if link and (link.startswith("http") or "drive.google.com" in link):
                match = await process_remote_resume(link, name, job_description)
                if match:
                    results.append({"filename": name, **match})
                    continue
            
            row_text = " ".join([str(c) for c in row if c is not None and str(c).strip()])
            if row_text.strip():
                results.append({"filename": name, **score_match(row_text, job_description)})
                
    except Exception:
        pass

@app.post("/match-excel-link")
async def match_excel_link(excel_link: str = Form(...), job_description: str = Form(...), top_n: int = Form(default=None)):
    try:
        url = get_drive_direct_link(excel_link)
        resp = requests.get(url, headers=HEADERS, timeout=30)
        
        if resp.status_code != 200:
            return JSONResponse({"error": "Could not download file from the provided link."}, 400)
            
        if b"<!DOCTYPE html>" in resp.content[:200]:
            return JSONResponse({"error": "Access Denied. Please make the Drive file 'Anyone with the link can view'."}, 400)

        results = []
        content_type = resp.headers.get("Content-Type", "").lower()
        is_excel = ("excel" in content_type or "spreadsheet" in content_type or 
                   "xlsx" in url.lower() or "xls" in url.lower())
        
        if is_excel:
            try:
                wb = load_workbook(io.BytesIO(resp.content), data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return {"results": []}
                
                headers = [str(c).lower().strip() if c else "" for c in rows[0]]
                name_idx, link_idx = 0, 1
                for i, h in enumerate(headers):
                    if any(k in h for k in ["name", "candidate", "full"]): name_idx = i
                    if any(k in h for k in ["link", "resume", "drive", "url"]): link_idx = i
                
                for i, row in enumerate(rows[1:], 1):
                    if not any(row): continue
                    name = str(row[name_idx]) if name_idx < len(row) and row[name_idx] else f"Candidate_{i}"
                    link = str(row[link_idx]) if link_idx < len(row) and row[link_idx] else ""
                    
                    if link and (link.startswith("http") or "drive.google.com" in link):
                        match = await process_remote_resume(link, name, job_description)
                        if match:
                            results.append({"filename": name, **match})
                            continue
                    
                    row_text = " ".join([str(c) for c in row if c is not None and str(c).strip()])
                    if row_text.strip():
                        match_result = score_match(row_text, job_description)
                        results.append({"filename": name, **match_result})
                        
            except Exception as e:
                return JSONResponse({"error": f"Failed to process Excel file: {str(e)}"}, 400)
        else:
            ext = ".pdf"
            if "word" in content_type or ".docx" in url.lower() or ".doc" in url.lower():
                ext = ".docx"
            elif "text" in content_type or ".txt" in url.lower():
                ext = ".txt"
            elif "pdf" in content_type or ".pdf" in url.lower():
                ext = ".pdf"
            
            text = extract_text(resp.content, f"resume{ext}")
            if text:
                result = score_match(text, job_description)
                results.append({"filename": "Resume_from_Drive", **result})
            else:
                return JSONResponse({"error": "Could not extract text from the resume file."}, 400)

        results = apply_dynamic_statuses(results, top_n)
        return {"results": results}
    except Exception as e:
        return JSONResponse({"error": f"Drive Link Error: {str(e)}"}, 500)

@app.post("/generate-excel")
async def generate_excel(data: AnalysisData):
    wb = Workbook(); ws = wb.active; ws.append(["Name", "Score (%)", "Status", "Matched Skills", "Missing Skills"])
    for r in data.results:
        ws.append([r.filename, r.score, r.status, ", ".join(r.matched_keywords), ", ".join(r.missing_keywords)])
    buf = io.BytesIO(); wb.save(buf)
    return Response(content=buf.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=Analysis_Report.xlsx"})

@app.get("/health")
async def health(): return {"status": "ok"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
